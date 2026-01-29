#!/usr/bin/env python3
"""
Generate citation metrics report for all authors from cited papers.
Saves progress incrementally in case of interruption.
"""

import json
import time
import requests
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from citation_metrics.config import get_api_key
from citation_metrics.metrics import compute_all_metrics


class OpenAlexClient:
    BASE_URL = "https://api.openalex.org"
    MAX_RETRIES = 5

    def __init__(self, api_key):
        self.api_key = api_key
        self.session = requests.Session()

    def _request(self, endpoint, params=None):
        if params is None:
            params = {}
        params["api_key"] = self.api_key
        url = f"{self.BASE_URL}{endpoint}"

        for attempt in range(self.MAX_RETRIES):
            try:
                resp = self.session.get(url, params=params, timeout=30)
                if resp.status_code == 429:
                    wait = 2 ** attempt
                    print(f"    Rate limited, waiting {wait}s...")
                    time.sleep(wait)
                    continue
                if resp.status_code == 404:
                    return None
                resp.raise_for_status()
                return resp.json()
            except requests.exceptions.RequestException as e:
                if attempt < self.MAX_RETRIES - 1:
                    time.sleep(2 ** attempt)
                    continue
                raise
        return None

    def get_author_by_orcid(self, orcid):
        """Fetch author metadata by ORCID."""
        data = self._request(f"/authors/orcid:{orcid}")
        if not data:
            return None
        return {
            "id": data["id"].split("/")[-1],
            "display_name": data.get("display_name", "Unknown"),
            "works_count": data.get("works_count", 0),
            "h_index_openalex": data.get("summary_stats", {}).get("h_index"),
            "orcid": orcid,
        }

    def get_works(self, author_id):
        """Fetch all works for an author."""
        works = []
        cursor = "*"

        while cursor:
            data = self._request("/works", params={
                "filter": f"author.id:{author_id}",
                "per_page": "200",
                "cursor": cursor,
                "select": "id,cited_by_count,authorships",
            })
            if not data:
                break

            for work in data.get("results", []):
                works.append({
                    "cited_by_count": work.get("cited_by_count", 0),
                    "num_authors": len(work.get("authorships", [])),
                })

            meta = data.get("meta", {})
            cursor = meta.get("next_cursor")

        return works


def load_orcids(filepath):
    """Load ORCID IDs from the output file."""
    orcids = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split(" | ")
            if len(parts) >= 3:
                orcid = parts[0].strip()
                openalex_id = parts[1].strip()
                name = parts[2].strip()
                orcids.append((orcid, openalex_id, name))
    return orcids


def load_progress(filepath):
    """Load previously computed results."""
    if not Path(filepath).exists():
        return {}
    with open(filepath) as f:
        return json.load(f)


def save_progress(filepath, results):
    """Save results incrementally."""
    with open(filepath, "w") as f:
        json.dump(results, f)


def create_xlsx(results, output_path):
    """Create xlsx file from results."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Author Metrics"

    # Headers
    headers = [
        "ORCID", "OpenAlex ID", "Name", "Papers", "Fractional Papers",
        "H-index", "H-index (OpenAlex)", "H/papers",
        "Weighted H", "Weighted H/frac papers",
        "Cit-divided H", "Cit-divided H/papers"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    row = 2
    for orcid, data in sorted(results.items(), key=lambda x: x[1].get("h_index", 0), reverse=True):
        if "error" in data:
            continue
        ws.cell(row=row, column=1, value=orcid)
        ws.cell(row=row, column=2, value=data.get("openalex_id", ""))
        ws.cell(row=row, column=3, value=data.get("name", ""))
        ws.cell(row=row, column=4, value=data.get("num_papers", 0))
        ws.cell(row=row, column=5, value=round(data.get("fractional_papers", 0), 2))
        ws.cell(row=row, column=6, value=data.get("h_index", 0))
        ws.cell(row=row, column=7, value=data.get("h_index_openalex", ""))
        ws.cell(row=row, column=8, value=round(data.get("h_per_paper", 0), 3))
        ws.cell(row=row, column=9, value=data.get("weighted_h", 0))
        ws.cell(row=row, column=10, value=round(data.get("weighted_h_per_frac_paper", 0), 3))
        ws.cell(row=row, column=11, value=data.get("citdiv_h", 0))
        ws.cell(row=row, column=12, value=round(data.get("citdiv_h_per_paper", 0), 3))
        row += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 30
    for col in "DEFGHIJKL":
        ws.column_dimensions[col].width = 12

    wb.save(output_path)
    return row - 2  # number of data rows


def main():
    input_file = "cited_author_orcids.txt"
    progress_file = "author_metrics_progress.json"
    output_file = "cited_authors_metrics.xlsx"

    print("Loading ORCID list...")
    authors = load_orcids(input_file)
    print(f"Found {len(authors)} authors to process")

    print("Loading previous progress...")
    results = load_progress(progress_file)
    print(f"Already processed: {len(results)} authors")

    api_key = get_api_key()
    client = OpenAlexClient(api_key)

    remaining = [(o, a, n) for o, a, n in authors if o not in results]
    print(f"Remaining to process: {len(remaining)} authors\n")

    for i, (orcid, openalex_id, name) in enumerate(remaining):
        print(f"[{i+1}/{len(remaining)}] Processing {name} ({orcid})...")

        try:
            # Get author info (verify they exist)
            author = client.get_author_by_orcid(orcid)
            if not author:
                results[orcid] = {"error": "Author not found", "name": name}
                save_progress(progress_file, results)
                continue

            # Get their works
            works = client.get_works(author["id"])
            if not works:
                results[orcid] = {
                    "error": "No works found",
                    "name": author["display_name"],
                    "openalex_id": author["id"],
                    "h_index_openalex": author["h_index_openalex"],
                }
                save_progress(progress_file, results)
                continue

            # Compute metrics
            metrics = compute_all_metrics(works)

            results[orcid] = {
                "openalex_id": author["id"],
                "name": author["display_name"],
                "h_index_openalex": author["h_index_openalex"],
                **metrics,
            }

            # Save progress every 10 authors
            if (i + 1) % 10 == 0:
                save_progress(progress_file, results)
                print(f"  Progress saved ({len(results)} total)")

        except Exception as e:
            print(f"  Error: {e}")
            results[orcid] = {"error": str(e), "name": name}
            save_progress(progress_file, results)

        # Small delay to be nice to the API
        time.sleep(0.1)

    # Final save
    save_progress(progress_file, results)

    # Generate xlsx
    print(f"\nGenerating {output_file}...")
    num_rows = create_xlsx(results, output_file)
    print(f"Done! Created {output_file} with {num_rows} authors")

    # Summary stats
    successful = sum(1 for r in results.values() if "error" not in r)
    errors = sum(1 for r in results.values() if "error" in r)
    print(f"\nSummary: {successful} successful, {errors} errors")


if __name__ == "__main__":
    main()
